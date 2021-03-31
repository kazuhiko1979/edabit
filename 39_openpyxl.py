# openpyxl

import openpyxl
import pprint


def main():

    workbook = openpyxl.load_workbook("Some.xlsx")
    sheet    = workbook["Sheet1"]

    # print(sheet["C5"].value)
    # print(sheet["C6"].value)
    # print(sheet["C7"].value)
    #
    # for i in sheet["C4:F8"][3]:
    #     print(i.value)

    # rows = sheet.iter_rows(
    #     min_row=4, max_row=8, min_col=3, max_col=6)
    # # print(rows)
    # # for i in list(rows)[2]:
    # #     print(i.value)
    #
    # pprint.pprint(list(sheet.values))

    sheet["C3"].value = "This is TEST Table."
    sheet.cell(row=2, column=3, value=100)
    workbook.save("Some.xlsx")
    # pprint.pprint(list(sheet.values))

    return

if __name__ == '__main__':
    main()

